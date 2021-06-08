
import openpyxl
import cv2
import pytesseract
from tkinter import *
'''
pytesseract.pytesseract.tesseract_cmd = r'/Library/Frameworks/Python.framework/Versions/3.9/bin'
img = cv2.imread('photo.png')
# to see the image below codes are there
cv2.imshow('sampleimage',img)
#enter any key to destroy the image window opened due to previous line code
cv2.waitKey(1)
cv2.destroyAllWindows()
#convert image to text using tesseract
text = pytesseract.image_to_string(img)
print(text)
'''


#Открытие файла с себестоимотсью
wb_obj = openpyxl.load_workbook("cena.xlsx")
sheet_obj = wb_obj.active
#Открытие файла с фактической стоимостью
wb_obj1 = openpyxl.load_workbook("fact.xlsx")
sheet_obj1 = wb_obj1.active
# списки для определения категорий
napitki = ['cola','fanta','water1','water2','water3','coffee1','coffee2','coffee3','sprite', 'вода', 'кола', 'сок']
food = ['sandwich','cookie','laysbig','layssmall','чипсы','сэндвич','печенье']
coffee = ['капучино','латте','раф']
cola = [ 'cola' , 'coke' , 'cocacola','кола','кокакола','колаж/б0,5']
i=1
j=1
n=(sheet_obj.max_row) 
b=(sheet_obj1.max_row)
n = n + 1
b = b + 1
row = 0
window = Tk()
window.title("Значение наценки")
#lbl = Label(window, text=('Добрый день. Введите максимально допустимое значение наценки '), font=("Arial", 24))
#lbl.grid(column=0, row=row)
row = row + 1

# спросим у пользователя значение наценки которое он считает критичным
'''
krit = Entry(window, width = 10)
krit.grid(column=1, row=0)
kritik = krit.get()
'''
fff = input('fff')
big=float(fff)
i=1
j=1
n=(sheet_obj.max_row) 
b=(sheet_obj1.max_row)
n = n + 1
b = b + 1

srednap=0
ssrednap = 0
sch = 0

sredfood = 0
ssredfood = 0
schfood = 0

sred_coffee = 0
ssred_coffee = 0
sch_coffee = 0


# Определение наценки и выявление нарушений        
while i < n and j<b:
    # товар с файла с сс
    tovar = sheet_obj.cell(row=j, column = 1)
    # товар с файла с фс
    tovar11 = sheet_obj1.cell(row=i, column = 1)
    
    cenaff = sheet_obj1.cell(row=i, column = 2)
    cenass = sheet_obj.cell(row=j, column = 2)
    if tovar11.value == tovar.value:
        k = cenaff.value
        l = cenass.value
        d = (k/l-1)*100
        if d > big:
            lbl = Label(window, text='наценка на товар ' + str(tovar.value) + ' равна ' + str(d))
            lbl.grid(column=0, row=row)
            row = row + 1
            #print('наценка на товар ' + str(tovar.value) + ' равна ' + str(d) )
        if tovar.value in napitki:
            srednap = srednap + d
            sch = sch + 1
            i = 1
            j = j + 1
        elif tovar.value in food:
             sredfood = sredfood + d
             schfood = schfood + 1
             i = 1
             j = j + 1
        elif tovar.value in coffee:
             sred_coffee = sred_coffee + d
             sch_coffee = sch_coffee + 1
             i = 1
             j = j + 1
        else:
            i = 1
            j = j + 1
    else:
        i = i + 1
ssrednap = srednap / sch
ssredfood = sredfood / schfood
ssred_coffee = sred_coffee / sch_coffee
def clicked():
    row1 = 0
    window = Tk()
    window.title("Значение наценки")
    lbl = Label(window, text=('средняя наценка на напитки составляет ' + str(ssrednap)), font=("Arial", 24))
    lbl.grid(column=0, row=row1)
    row1 = row1 + 1
    lbl = Label(window, text=('средняя наценка на еду составляет ' + str(ssredfood)), font=("Arial", 24))
    lbl.grid(column=0, row=row1)
    row1 = row1 + 1
    lbl = Label(window, text=('средняя наценка на кофе составляет ' + str(ssred_coffee)), font=("Arial", 24))
    lbl.grid(column=0, row=row1)
    row1 = row1 + 1
    '''
    res = "Привет {}".format(krit.get())
    lbl = Label(window, text=(res), font=("Arial", 24))
    lbl.grid(column=0, row=row1)
    row1 = row1 + 1
    lbl = Label(window, text=(kritik), font=("Arial", 24))
    lbl.grid(column=0, row=row1)
    row1 = row1 + 1
    '''
btn = Button(window, text="рассчитать среднее по категориям", font=("Arial Bold", 24), command=clicked)
btn.grid(column=0, row=row)
row = row + 1
#print('средняя наценка на напитки составляет ' + str(ssrednap))
#print('средняя наценка на еду составляет ' + str(ssredfood))
#print('средняя наценка на кофе составляет ' + str(ssred_coffee))







